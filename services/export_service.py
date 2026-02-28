"""
Export Service — builds XLSX exports for applicant data.
No Flask. Returns BytesIO for the API layer to serve via send_file / make_response.
"""

import io
import json
import re
from datetime import datetime
from models.applicants import (
    get_all_applicants_complete_export,
    get_selected_applicants_for_export,
)
from utils.csv_helpers import clean_row
from utils.activity_logger import log_activity

# Strip control characters illegal in XLSX/XML (keeps tab, newline, carriage return)
_ILLEGAL_XML_CHARS = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


def _build_xlsx(rows: list) -> io.BytesIO:
    """Convert a list of dicts to a formatted XLSX BytesIO."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Applicants"

    headers = list(rows[0].keys())
    ws.append(headers)

    for row_data in rows:
        cleaned = clean_row(row_data)
        row = []
        for h in headers:
            val = cleaned.get(h, "")
            if isinstance(val, (list, dict)):
                val = json.dumps(val)
            if isinstance(val, str):
                val = _ILLEGAL_XML_CHARS.sub(", ", val)
            row.append(val)
        ws.append(row)

    # Force all cells to text format to prevent Excel date casting
    for col_idx in range(1, len(headers) + 1):
        for row_idx in range(1, ws.max_row + 1):
            ws.cell(row=row_idx, column=col_idx).number_format = "@"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


class ExportService:

    def export_all(self, exporter_email: str, session_id=None) -> tuple:
        """
        Export all applicants as XLSX.
        Returns (BytesIO, filename).
        Raises ValueError if no applicants found.
        """
        applicants, error = get_all_applicants_complete_export()
        if error:
            raise ValueError(error)
        if not applicants:
            raise ValueError("No applicants found in database")

        output = _build_xlsx(applicants)

        log_activity(
            action_type="export",
            target_entity="all_applicants",
            target_id="complete_database",
            additional_metadata={
                "record_count": len(applicants),
                "export_type": "complete_database_export",
                "exported_by": exporter_email,
            },
        )

        filename = f"complete_export_{len(applicants)}_applicants_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
        return output, filename

    def export_selected(self, user_codes: list, sections: list | None, exporter_email: str) -> tuple:
        """
        Export selected applicants as XLSX.
        Returns (BytesIO, filename).
        Raises ValueError on error.
        """
        if not user_codes:
            raise ValueError("No applicants selected")

        applicants, error = get_selected_applicants_for_export(user_codes, sections)
        if error:
            raise ValueError(error)
        if not applicants:
            raise ValueError("No applicant data found for selected users")

        output = _build_xlsx(applicants)

        log_activity(
            action_type="export",
            target_entity="selected_applicants",
            target_id=None,
            additional_metadata={
                "record_count": len(applicants),
                "user_codes": user_codes,
                "sections": sections or "all",
                "export_style": "horizontal_dynamic_pivoted",
            },
        )

        sections_str = "_".join(sections) if sections else "all"
        filename = f"selected_{len(user_codes)}_{sections_str}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return output, filename
